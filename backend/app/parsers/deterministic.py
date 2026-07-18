from __future__ import annotations

import re
from io import BytesIO
from zipfile import BadZipFile, ZipFile
from dataclasses import dataclass
from xml.etree import ElementTree

from openpyxl import load_workbook
from pypdf import PdfReader


@dataclass(frozen=True)
class ParsedPage:
    page_number: int
    raw_text: str
    layout_json: dict
    ocr_used: bool = False


class DeterministicTextParser:
    """Safe demo adapter; never executes document instructions or embedded code."""

    page_marker = re.compile(r"(?:^|\n)\s*(?:={2,}\s*)?第\s*\d+\s*页(?:\s*={2,})?\s*(?:\n|$)")

    def parse(self, data: bytes, mime_type: str) -> list[ParsedPage]:
        if data.startswith(b"%PDF") and mime_type == "application/pdf":
            return self._parse_pdf(data)
        if mime_type.startswith("image/"):
            return [
                ParsedPage(
                    1,
                    "",
                    {
                        "adapter": "image-ocr-placeholder-v1",
                        "ocr_required": True,
                        "blocks": [],
                    },
                )
            ]
        if mime_type.endswith("wordprocessingml.document"):
            return self._parse_docx(data)
        if mime_type.endswith("spreadsheetml.sheet"):
            return self._parse_xlsx(data)
        payload = data.split(b"\n", 1)[1] if data.startswith(b"%PDF") else data
        text = payload.decode("utf-8", errors="replace").replace("\x00", "")
        chunks = text.split("\f")
        if len(chunks) == 1:
            chunks = self.page_marker.split(text)
        chunks = [item.strip() for item in chunks if item.strip()]
        if not chunks:
            chunks = ["[该页未提取到可用文本，需人工复核/OCR]"]
        return [
            ParsedPage(
                page_number=index,
                raw_text=chunk,
                layout_json={
                    "adapter": "deterministic-text-v1",
                    "blocks": [{"type": "paragraph", "bbox": [0, 0, 1, 1], "text": chunk}],
                },
            )
            for index, chunk in enumerate(chunks, start=1)
        ]

    def _parse_pdf(self, data: bytes) -> list[ParsedPage]:
        reader = PdfReader(BytesIO(data), strict=False)
        pages = []
        for index, pdf_page in enumerate(reader.pages, start=1):
            text = (pdf_page.extract_text() or "").strip()
            ocr_required = not text
            pages.append(
                ParsedPage(
                    index,
                    text,
                    {
                        "adapter": "pypdf-v1",
                        "ocr_required": ocr_required,
                        "blocks": [{"type": "page_text", "bbox": [0, 0, 1, 1], "text": text}],
                    },
                )
            )
        if not pages:
            raise ValueError("PDF contains no pages")
        return pages

    def _parse_docx(self, data: bytes) -> list[ParsedPage]:
        try:
            with ZipFile(BytesIO(data)) as archive:
                infos = archive.infolist()
                if len(infos) > 2000 or sum(item.file_size for item in infos) > 50 * 1024 * 1024:
                    raise ValueError("DOCX archive exceeds safe expansion limits")
                if any(".." in item.filename.split("/") for item in infos):
                    raise ValueError("DOCX contains unsafe paths")
                xml_data = archive.read("word/document.xml")
        except (BadZipFile, KeyError) as exc:
            raise ValueError("invalid DOCX package") from exc
        root = ElementTree.fromstring(xml_data)
        namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
        body = root.find(f"{namespace}body")
        if body is None:
            raise ValueError("DOCX document body is missing")
        raw_lines = []
        blocks = []
        tables = []
        paragraph_index = 0
        table_index = 0
        for child in body:
            if child.tag == f"{namespace}p":
                text = "".join(node.text or "" for node in child.iter(f"{namespace}t")).strip()
                if not text:
                    continue
                paragraph_index += 1
                raw_lines.append(text)
                blocks.append(
                    {"type": "paragraph", "paragraph_index": paragraph_index, "text": text}
                )
            elif child.tag == f"{namespace}tbl":
                table_index += 1
                rows = []
                for row_index, row in enumerate(child.findall(f"{namespace}tr"), start=1):
                    cells = []
                    for cell_index, cell in enumerate(row.findall(f"{namespace}tc"), start=1):
                        value = "".join(
                            node.text or "" for node in cell.iter(f"{namespace}t")
                        ).strip()
                        cells.append({"cell": f"R{row_index}C{cell_index}", "text": value})
                    rows.append(cells)
                    row_text = " | ".join(cell["text"] for cell in cells if cell["text"])
                    if row_text:
                        raw_lines.append(row_text)
                table_data = {"table_index": table_index, "rows": rows}
                tables.append(table_data)
                blocks.append({"type": "table", **table_data})
        return [
            ParsedPage(
                1,
                "\n".join(raw_lines) or "[文档未提取到可用文本，需人工复核]",
                {
                    "adapter": "docx-xml-v2",
                    "blocks": blocks,
                    "tables": tables,
                    "logical_location": "body_order",
                },
            )
        ]

    def _parse_xlsx(self, data: bytes) -> list[ParsedPage]:
        try:
            workbook = load_workbook(BytesIO(data), read_only=True, data_only=False)
            values_workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError("invalid XLSX package") from exc
        pages = []
        for index, sheet in enumerate(workbook.worksheets, start=1):
            values_sheet = values_workbook[sheet.title]
            cells = []
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    formula = cell.data_type == "f"
                    cached_value = values_sheet[cell.coordinate].value if formula else None
                    cells.append(
                        {
                            "address": cell.coordinate,
                            "value": str(cell.value),
                            "formula": formula,
                            "cached_value": str(cached_value) if cached_value is not None else None,
                            "needs_review": formula and cached_value is None,
                        }
                    )
            raw_text = "\n".join(
                f"{cell['address']}: {cell['value']}"
                + (
                    f" => {cell['cached_value']}"
                    if cell["formula"] and cell["cached_value"] is not None
                    else " [公式缓存值缺失，需人工复核]"
                    if cell["needs_review"]
                    else ""
                )
                for cell in cells
            )
            pages.append(ParsedPage(index, raw_text, {"adapter": "openpyxl-v1", "sheet": sheet.title, "cells": cells}))
        workbook.close()
        values_workbook.close()
        if not pages:
            raise ValueError("XLSX contains no sheets")
        return pages
